# -*- encoding: utf-8 -*-
"""
@时间:   2021/12/25 10:57
@作者:   王齐涛
@文件:   excel_openyxl_data.py
"""
import json
from openpyxl import load_workbook


class HandleExcel:
    """使用内置库openyxl，进行Excel的封装操作,暂时没有用到该方法"""

    __workbook = None  # __workbook用于存放excel文件的对象
    __sheet = None
    __print = True  # 用于开启全局打印

    def __init__(self, file_name, sheet_name):
        if file_name:
            self.file_name = file_name
        if sheet_name:
            self.sheet_name = sheet_name
        HandleExcel.__workbook = load_workbook(filename=self.file_name)  # 实例化对象
        HandleExcel.__sheet = HandleExcel.__workbook[self.sheet_name]   # 获取sheet

    def get_test_case(self):
        """
        获取excel表的测试数据
        :return:
        """
        # 列表
        all_excel_data = list(self.__sheet.iter_rows(values_only=True))  # 获取表格所有数据
        excel_title = all_excel_data[0]  # 获取表头
        case_data_list = all_excel_data[1:]  # 获取所有的用例数据
        test_case_list = []  # 新建空列表，收集所有的测试用例数据
        for case in case_data_list:
            test_case = dict(zip(excel_title, case))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
            # print(test_case)
            test_case_list.append(test_case)  # 将每次拼接好的测试用例添加到test_case_list
        self.__workbook.close()   # 读取完进行关闭
        if self.__print is True:
            print(f"获取到的数据显示{test_case_list}")

        # 封装成字典
        # all_excel_data = list(self.__sheet.iter_rows(values_only=True))  # 获取表格所有数据
        # excel_title = all_excel_data[0]  # 获取表头
        # case_data_list = all_excel_data[1:]  # 获取所有的用例数据
        # test_case_list = {}  # 新建空列表，收集所有的测试用例数据
        # for case in case_data_list:
        #     test_case = dict(zip(excel_title, case))  # 数据拼接，用表头与测试用例数据进行组装拼接成dict
        #     print(f"获取到的数据显示{test_case}")

        return test_case_list

    def write_test_data(self, row, column, value):
        """
        往Excel表中写入数据
        :param row: 行
        :param column: 列
        :param value: 写入的值
        :return:
        """
        self.__sheet.cell(row=row, column=column, value=value)
        self.__workbook.save(self.file_name)
        print(f"写入数据： {value} 到Excel中成功！")

    def get_max_row(self):
        """
        获取Excel表中的最大行数
        :return:
        """
        max_row = int(self.__sheet.max_row)
        if self.__print is True:
            print(f"最大行数{max_row}")
        return max_row

    def get_max_col(self):
        """
        获取Excel表中的最大列数
        :return:
        """
        max_col = int(self.__sheet.max_row)
        if self.__print is True:
            print(f"最大行数{max_col}")
        return max_col

    def get_rows_value(self, row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_test_case()[row]:
            row_list.append(i.value)
        return row_list

    def get_row(self, search_value):
        # 初始化行数
        row_number = None
        # 遍历工作表中的每一行
        for row in self.__sheet.iter_rows():
            for cell in row:
                if cell.value == search_value:
                    row_number = cell.row
                    break
            if row_number is not None:
                break
        # 如果找到数据，打印行数
        if row_number is not None:
            print(f'数据 "{search_value}" 位于第 {row_number} 行')
        else:
            print(f'在工作表中未找到数据 "{search_value}"')
        return row_number


def replace_line_by_search_string(file_path, search_string, replace_string):
    # 读取原始文本文件内容
    with open(file_path, 'r', encoding='gbk', errors='ignore') as file:
        lines = file.readlines()
    # 遍历每一行，判断是否包含搜索字符串
    for i in range(len(lines)):
        if search_string in lines[i]:
            lines[i] = replace_string + '\n'  # 替换整行内容
    # 将修改后的文本内容写入到原始文本文件中
    with open(file_path, 'w', encoding='gbk', errors='ignore') as file:
        file.writelines(lines)
    print(f'搜索字符串 "{search_string}" ,已成功替换为 "{replace_string}"')


def replace_key(excel_path, js_path):
    list_data = HandleExcel(excel_path, "Sheet1").get_test_case()
    print(list_data)
    with open(js_path, 'r', encoding='gbk', errors='ignore') as file:
        a = file.read()
        for i in list_data:
            if i['name'] in a:
                search_string = i['name']  # 搜索字符串
                replace_string = " "+"'"+i['name']+"':'"+i['tihuan']+"'"    # 替换字符串样式
                replace_line_by_search_string(js_path, search_string, replace_string)
            else:
                print(f"excel中的字段：{i['name']}在文档中没有找到")


if __name__ == '__main__':
    HandleExcel(r"N:\\UI_WEB_ALL\\data\\ACBN.xlsx", "Sheet1").get_test_case()
    # replace_key(r"N:\AB.xlsx", r"N:\test11.txt")





